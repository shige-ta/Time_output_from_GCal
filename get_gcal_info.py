# -*- coding: utf-8 -*-
import datetime
import pickle
import os.path
import openpyxl
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import configparser

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/calendar.readonly']


def main():
    # コンフィグファイル読み込み
    config_ini = configparser.ConfigParser()
    config_ini.read('config.ini', encoding='utf-8')
    calendar_id = config_ini['DEFAULT']['calendar_id']

    creds = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('calendar', 'v3', credentials=creds)

    # Call the Calendar API
    # TODO maxResultsの上限 startとendの設定 
    now = datetime.datetime.utcnow().isoformat() + 'Z' # 'Z' indicates UTC time
    events_result = service.events().list(calendarId=calendar_id, timeMin=now,
                                        maxResults=10, singleEvents=True,).execute()
    events = events_result.get('items', [])

    #エクセル書き込み
    if not events:
        print('イベントゼロ.')
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, event in enumerate(events):
        schedule_start = event['start'].get('dateTime', event['start'].get('date'))
        schedule_end = event['end'].get('dateTime', event['end'].get('date'))
        ws['A' + str(i+1)].value = str(schedule_start)
        ws['B' + str(i+1)].value = str(schedule_end)

    wb.save('output.xlsx')

if __name__ == '__main__':
    main()