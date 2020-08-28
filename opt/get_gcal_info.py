#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# Copyright 2014 Google Inc. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Simple command-line sample for the Calendar API.
Command-line application that retrieves the list of the user's calendars."""

import sys
import datetime
import os.path
import openpyxl
import configparser
from oauth2client import client
from googleapiclient import sample_tools


def main(argv):
    # コンフィグファイル読み込み
    config_ini = configparser.ConfigParser()
    config_ini.read('config.ini', encoding='utf-8')
    calendar_id = config_ini['DEFAULT']['calendar_id']

    # Authenticate and construct service.
    service, flags = sample_tools.init(
        argv, 'calendar', 'v3', __doc__, __file__,
        scope='https://www.googleapis.com/auth/calendar.readonly')

    # Call the Calendar API
    # TODO maxResultsの上限 startとendの設定 
    now = datetime.datetime.utcnow().isoformat() + 'Z' # 'Z' indicates UTC time
    events_result = service.events().list(calendarId=calendar_id, timeMin=now,
                                        maxResults=10, singleEvents=True,).execute()
    events = events_result.get('items', [])

    #エクセル書き込み
    if not events:
        print('イベントゼロ.')
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, event in enumerate(events):
        schedule_start = event['start'].get('dateTime', event['start'].get('date'))
        schedule_end = event['end'].get('dateTime', event['end'].get('date'))
        ws['A' + str(i+1)].value = str(schedule_start)
        ws['B' + str(i+1)].value = str(schedule_end)

    wb.save('output.xlsx')
if __name__ == '__main__':
    main(sys.argv)
